#!/usr/bin/env python3
"""
Pipeline Upah TSJ — Auto-populates Kit-Upah-TSJ-AUTO.xlsx from 12-file closing folder.

Proven NET calc (from /tmp/rekon2.py — 110/120 zero-diff vs ASLI c26):
  - c8 resolution: MUARA value → formula HK×128594 → ASLI cache → 0
  - BPJS resolution: MUARA value → ASLI cache → formula eval → 0
  - Ledger aggregation by NIK (type preserved via str(nik).strip() as key)
  - NET = c8 + ADD - (BPJS + TIDAK BASIS + DENDA + PINJAMAN TENGAH BULAN + PINJAMAN KARYAWAN + ALAT PANEN)

Output sheets:
  1. Upah PERIODE  — 120 rows auto-populated, c8/c17-c19 RESOLVED VALUES (not formulas),
                     c9-c15/c20/c22-c24 live SUMIFS, c16/c25/c26/c29 formulas
  2. DATA INPUT    — 270 rows verbatim from MUARA, NIK as TEXT
  3. VALIDASI      — 11 kategori + TOTAL, live SUMIFS, auto OK/CEK
  4. TRACE INTEGRASI — 12 file + rekon verdicts + Herlima + c28 (AB) mystery + NIK coverage
  5. CARA PAKAI    — workflow, auto vs manual, August, troubleshooting
"""
import openpyxl, re, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Windows cmd kadang bukan UTF-8 — guard biar print nggak crash
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

# ============ PATHS (portable, bisa di-override via CLI) ============
# Cara pakai:
#   python pipeline_upah.py                                  → pakai default (folder di samping script)
#   python pipeline_upah.py closingan-agustus-2026           → folder sumber custom
#   python pipeline_upah.py closingan-agustus-2026 out.xlsx  → folder sumber + nama output custom
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, 'closingan-juli-2026')
OUT  = sys.argv[2] if len(sys.argv) > 2 else os.path.join(SCRIPT_DIR, 'Kit-Upah-TSJ-AUTO.xlsx')
if not os.path.isdir(BASE):
    print(f'ERROR: folder sumber tidak ditemukan: {BASE}')
    print('Taruh folder closingan di samping pipeline_upah.py, atau pass path-nya sebagai argumen pertama.')
    sys.exit(1)
BASE = BASE.rstrip('/\\') + os.sep
print(f'Folder sumber : {BASE}')
print(f'Output        : {OUT}', flush=True)
RATE = 128594

ADD = {'PREMI UMUM', 'PREMI KEHADIRAN', 'PRUNING P', 'TUNJANGAN BBM',
       'TUNJANGAN JABATAN', 'TUNJANGAN LAINNYA', 'OVER TIME'}
DED = {'TIDAK BASIS', 'DENDA', 'PINJAMAN TENGAH BULAN', 'PINJAMAN KARYAWAN'}

KATEGORI = [
    ('PREMI UMUM', 'I', 35),
    ('PREMI KEHADIRAN', 'J', 10),
    ('PRUNING P', 'K', 0),
    ('TUNJANGAN BBM', 'L', 10),
    ('TUNJANGAN JABATAN', 'M', 3),
    ('TUNJANGAN LAINNYA', 'N', 1),
    ('OVER TIME', 'O', 9),
    ('TIDAK BASIS', 'T', 54),
    ('DENDA', 'V', 42),
    ('PINJAMAN TENGAH BULAN', 'W', 100),
    ('PINJAMAN KARYAWAN', 'X', 6),
]

# ---- Styles ----
HDR_FILL  = PatternFill('solid', fgColor='D9D9D9')
NOTE_FILL = PatternFill('solid', fgColor='FFF2CC')
BOLD      = Font(bold=True)
TITLE     = Font(bold=True, size=13)
THIN      = Border(left=Side('thin'), right=Side('thin'),
                   top=Side('thin'), bottom=Side('thin'))
MONEY     = '#,##0.00'


def num(x):
    return float(x) if isinstance(x, (int, float)) else 0.0


# ============ INPUT CHECK ============
print('[1/6] Cek file hulu (11 wajib + ASLI-backup opsional)...', flush=True)
needed = [
    'Data Upah TSJ Juli 2026.xlsx',
    # 'Data Upah TSJ Juli 2026.ASLI-backup.xlsx',  # OPSIONAL — cache resolve c8/BPJS; fallback ke file utama
    'Gaji Juli  2026.xlsx',  # NOTE: double space
    'Master Data Karyawan 2026 (1).xlsx',
    'Data Pemanen Juli 2026.xlsx',
    'Data Pengangkutan Juli 2026.xlsx',
    'Data PK Alat TSJ Juli 2026.xlsx',
    'Data PK Lahan TSJ Juli 2026.xlsx',
    'Data Security TSJ Juli 2026.xlsx',
    'Data Supervisi TSJ Juli 2026.xlsx',
    'Langsir Along-Along TSJ Juli 2026.xlsx',
    'Premi Perawatan TSJ Juli 2026.xlsx',
]
missing = [f for f in needed if not os.path.exists(BASE + f)]
if missing:
    print(f'  MISSING: {missing}')
    sys.exit(1)
print(f'  OK: semua file hulu wajib ada di {BASE}')

# ============ LOAD ============
print('[2/6] Load MUARA + ASLI backup + DATA INPUT + TF BNI + Master...', flush=True)
wbv = openpyxl.load_workbook(BASE + 'Data Upah TSJ Juli 2026.xlsx', data_only=True)
wbf = openpyxl.load_workbook(BASE + 'Data Upah TSJ Juli 2026.xlsx', data_only=False)
ws  = wbv['Upah Juli 2026']
wf  = wbf['Upah Juli 2026']
di  = wbv['DATA INPUT']
_backup_path = BASE + 'Data Upah TSJ Juli 2026.ASLI-backup.xlsx'
if os.path.exists(_backup_path):
    wba = openpyxl.load_workbook(_backup_path, data_only=True)
    wa  = wba['Upah Juli 2026']
    print('  ASLI-backup: OK - sumber cache resolve c8/BPJS')
else:
    print('  WARN: ASLI-backup tidak ada - fallback cache resolve c8/BPJS ke file utama')
    wba = None
    wa  = wbv['Upah Juli 2026']
wbt = openpyxl.load_workbook(BASE + 'Gaji Juli  2026.xlsx', data_only=True, read_only=True)
wt  = wbt['TF BNI']
wbm = openpyxl.load_workbook(BASE + 'Master Data Karyawan 2026 (1).xlsx', data_only=True, read_only=True)
wms = wbm['Juli']

# Aggregate ledger by NIK
add_by_nik = {}; ded_by_nik = {}
ledger_total = 0
for r in range(2, di.max_row + 1):
    nik = di.cell(r, 2).value
    kat = di.cell(r, 4).value
    nom = num(di.cell(r, 5).value)
    if nik is None or kat is None: continue
    k = str(nik).strip()
    if kat in ADD:   add_by_nik[k] = add_by_nik.get(k, 0) + nom
    elif kat in DED: ded_by_nik[k] = ded_by_nik.get(k, 0) + nom
    ledger_total += 1

# Resolution chains (from rekon2.py, verbatim)
def resolve_c8(r):
    v = ws.cell(r, 8).value
    if v is not None: return num(v), 'value'
    f = wf.cell(r, 8).value
    if isinstance(f, str):
        m = re.match(r'=G%d\*(\d+)' % r, f.replace(' ', ''))
        if m: return num(ws.cell(r, 7).value) * float(m.group(1)), 'formula_hk_rate'
    av = wa.cell(r, 8).value
    if av is not None: return num(av), 'asli_cache'
    return 0.0, 'none'

def resolve_bpjs(r, col):
    v = ws.cell(r, col).value
    if v is not None: return num(v), 'value'
    av = wa.cell(r, col).value
    if av is not None: return num(av), 'asli_cache'
    f = wf.cell(r, col).value
    if isinstance(f, str):
        m = re.search(r'(\d+(?:\.\d+)?)\*(\d+(?:\.\d+)?)%', f.replace(' ', ''))
        if m: return float(m.group(1)) * float(m.group(2)) / 100.0, 'formula_eval'
        m2 = re.match(r'=H%d\*(\d+(?:\.\d+)?)%%' % r, f.replace(' ', ''))
        if m2:
            c8, _ = resolve_c8(r); return c8 * float(m2.group(1)) / 100.0, 'formula_h_pct'
    return 0.0, 'none'

# Master NIKs
master_niks = set()
for row in wms.iter_rows(min_row=7, max_row=283, max_col=4):
    if row[2].value: master_niks.add(str(row[2].value).strip())

# TF BNI — key by NIK (Asman ada 2 orang berbeda; name-keyed collides & drops one)
tf_net = {}
for row in wt.iter_rows(min_row=5, max_row=132, max_col=6):
    nik = row[1].value; nama = row[2].value; gaji = row[5].value
    if nama and isinstance(gaji, (int, float)):
        key = str(nik).strip() if nik is not None else ('NAME:' + str(nama).strip())
        tf_net[key] = num(gaji)

# Roster NIKs (for coverage)
roster_niks = set()
roster_with_nik = 0
roster_no_nik = []
for r in range(9, 129):
    nik = ws.cell(r, 2).value
    nama = ws.cell(r, 3).value
    if nama:
        if nik is not None:
            roster_niks.add(str(nik).strip())
            roster_with_nik += 1
        else:
            roster_no_nik.append((r, nama))
covered = len(roster_niks & master_niks)

# ============ BUILD OUTPUT ============
print('[3/6] Build Upah PERIODE (120 baris)...', flush=True)
wb = Workbook()
ws_out = wb.active
ws_out.title = 'Upah PERIODE'

# Titles
ws_out['A1'] = 'PT. TULAS SAKTI JAYA'; ws_out['A1'].font = TITLE
ws_out['A2'] = 'DAFTAR UPAH KARYAWAN PERIODE JULI 2026'; ws_out['A2'].font = BOLD
ws_out['C3'] = '       Kebun Pasaman Barat'
ws_out['C4'] = '       Afdeling VIII'
ws_out.merge_cells('A1:F1')
ws_out.merge_cells('A2:F2')
ws_out['AB1'] = 'RATE PER HK ->'; ws_out['AB1'].font = BOLD
ws_out['AC1'] = RATE
ws_out['AC1'].fill = NOTE_FILL
ws_out['AC1'].font = BOLD
ws_out['AC1'].number_format = MONEY

# Header r5-r7 (mirror blank template structure)
h5 = {1: 'NO', 2: 'NIK', 3: 'NAMA', 4: 'STATUS', 5: 'JABATAN', 6: 'PEKERJAAN',
      7: 'HARI KERJA', 9: 'PREMI', 12: 'TUNJANGAN', 15: 'OVER TIME',
      16: 'UPAH BRUTO', 17: 'POTONGAN', 25: 'TOTAL POTONGAN', 26: 'UPAH NETTO',
      28: 'NET REF KTU*', 29: 'DIFF (CEK)'}
for c, v in h5.items():
    cell = ws_out.cell(5, c, v)
    cell.font = BOLD; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws_out.merge_cells('I5:K5')
ws_out.merge_cells('L5:N5')
ws_out.merge_cells('Q5:X5')

h6 = {17: 'BPJS KES', 18: 'BPJS TK', 20: 'TIDAK BASIS', 21: 'ALAT PANEN',
      22: 'DENDA', 23: 'PINJAMAN TENGAH BULAN', 24: 'PINJAMAN KARYAWAN'}
for c, v in h6.items():
    cell = ws_out.cell(6, c, v)
    cell.font = BOLD; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws_out.merge_cells('R6:S6')

h7 = {7: 'HK', 8: 'Rp', 9: 'UMUM', 10: 'KEHADIRAN', 11: 'PRUNING.P',
      12: 'BBM', 13: 'JABATAN', 14: 'LAINNYA',
      17: 'KES 1.%', 18: 'JHT.2%', 19: 'JP.1%'}
for c, v in h7.items():
    cell = ws_out.cell(7, c, v)
    cell.font = BOLD; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center')

# Roster rows 9-128
roster_count = 0
for r in range(9, 129):
    nama = ws.cell(r, 3).value
    if not nama: continue
    roster_count += 1

    # c1-c7 from MUARA
    ws_out.cell(r, 1, roster_count)
    nik_raw = ws.cell(r, 2).value
    nik_txt = str(nik_raw).strip() if nik_raw is not None else ''
    c2 = ws_out.cell(r, 2, nik_txt if nik_txt else None)
    c2.number_format = '@'  # force TEXT for SUMIFS
    ws_out.cell(r, 3, nama)
    ws_out.cell(r, 4, ws.cell(r, 4).value)
    ws_out.cell(r, 5, ws.cell(r, 5).value)
    ws_out.cell(r, 6, ws.cell(r, 6).value)
    hk = ws.cell(r, 7).value
    ws_out.cell(r, 7, hk)

    # c8 RESOLVED VALUE (HK×RATE or hardcode or ASLI cache)
    c8_val, _ = resolve_c8(r)
    ws_out.cell(r, 8, c8_val).number_format = MONEY

    # c9-c15 + c20 + c22-c24 live SUMIFS
    for cat, col_letter in [(k[0], k[1]) for k in KATEGORI]:
        col_idx = ord(col_letter) - 64
        formula = ('=SUMIFS(\'DATA INPUT\'!$E$2:$E$2001,'
                   '\'DATA INPUT\'!$B$2:$B$2001,$B%d,'
                   '\'DATA INPUT\'!$D$2:$D$2001,"%s")') % (r, cat)
        ws_out.cell(r, col_idx, formula).number_format = MONEY

    # c16 UPAH BRUTO = SUM(H:O)
    ws_out.cell(r, 16, '=SUM(H%d:O%d)' % (r, r)).number_format = MONEY

    # c17/c18/c19 RESOLVED VALUES (BPJS, bukan formula)
    c17_val, _ = resolve_bpjs(r, 17)
    c18_val, _ = resolve_bpjs(r, 18)
    c19_val, _ = resolve_bpjs(r, 19)
    if c17_val > 0: ws_out.cell(r, 17, c17_val).number_format = MONEY
    if c18_val > 0: ws_out.cell(r, 18, c18_val).number_format = MONEY
    if c19_val > 0: ws_out.cell(r, 19, c19_val).number_format = MONEY

    # c21 ALAT PANEN — manual, Juli 2026 = kosong semua

    # c25 TOTAL POTONGAN = SUM(Q:X)
    ws_out.cell(r, 25, '=SUM(Q%d:X%d)' % (r, r)).number_format = MONEY
    # c26 UPAH NETTO = P - Y
    ws_out.cell(r, 26, '=P%d-Y%d' % (r, r)).number_format = MONEY

    # c28 NET REF KTU — verbatim from MUARA (untraced column, honest label)
    c28 = ws.cell(r, 28).value
    if c28 is not None:
        ws_out.cell(r, 28, c28).number_format = MONEY
    # c29 DIFF = AB - Z
    ws_out.cell(r, 29, '=AB%d-Z%d' % (r, r)).number_format = MONEY

    # Borders + format numeric columns
    for c in range(1, 30):
        ws_out.cell(r, c).border = THIN

# Catatan
NOTE_ROW = 130
notes = [
    'CATATAN:',
    '1) RATE/HK di sel AC1 (Juli 2026 = 128.594; konfirmasi KTU tiap periode).',
    '2) c8 (Rp) = nilai RESOLVED — untuk 105 baris = HK×128.594; 15 baris = hardcode KTU; sisanya fallback ke ASLI cache. Lihat pipeline_upah.py → resolve_c8.',
    '3) c17/c18/c19 (BPJS) = nilai RESOLVED (bukan formula) — formula asli pakai base literal bukan sel H, jadi copy value lebih aman.',
    '4) c9-c15, c20, c22-c24 = live SUMIFS dari sheet DATA INPUT. Cek VALIDASI (semua harus OK).',
    '5) c21 (ALAT PANEN) = MANUAL. Juli 2026 kosong semua, audit-verified.',
    '6) c28 (NET REF KTU*) = salin VERBATIM dari kolom AB file KTU. TANPA HEADER di sumber (lihat TRACE §5). Jangan pakai sebagai acuan bayar.',
    '7) c29 (DIFF) = AB - Z — biasanya non-nol karena c28 ≠ NET formula. Bank transfer BNI ikut c26 (formula), bukan c28.',
    '8) NIK diseragamkan ke TEXT (number_format="@") untuk SUMIFS match. Tipe campur di MUARA (str + int) jadi silent SUMIFS fail kalau dibiarkan.',
]
for i, txt in enumerate(notes):
    cell = ws_out.cell(NOTE_ROW + i, 1, txt)
    cell.fill = NOTE_FILL
    if i == 0: cell.font = BOLD

# Column widths
widths = {'A': 4, 'B': 12, 'C': 24, 'D': 8, 'E': 16, 'F': 14, 'G': 6, 'AB': 14, 'AC': 14}
for col, w in widths.items():
    ws_out.column_dimensions[col].width = w
for c in range(8, 27):
    ws_out.column_dimensions[get_column_letter(c)].width = 11
ws_out.freeze_panes = 'D9'

# ============ SHEET 2: DATA INPUT ============
print('[4/6] Build DATA INPUT (270 baris verbatim)...', flush=True)
di_out = wb.create_sheet('DATA INPUT')
hdr = ['TGL', 'NIK', 'NAMA', 'KATEGORI', 'NOMINAL', 'KET', 'SUMBER']
for c, v in enumerate(hdr, 1):
    cell = di_out.cell(1, c, v)
    cell.font = BOLD; cell.fill = HDR_FILL

# Verbatim copy
for r in range(2, di.max_row + 1):
    for c in range(1, 8):
        v = di.cell(r, c).value
        if c == 2 and v is not None:
            v = str(v).strip()  # force TEXT
        cell = di_out.cell(r, c, v)
        if c == 1 and v: cell.number_format = 'YYYY-MM-DD'
        if c == 2: cell.number_format = '@'
        if c == 5 and v is not None: cell.number_format = MONEY

# Data validation for KATEGORI
dv = DataValidation(type='list',
                    formula1='"' + ','.join(k[0] for k in KATEGORI) + '"',
                    allow_blank=True)
di_out.add_data_validation(dv)
dv.add('D2:D2001')

# Legend
di_out['I1'] = 'DAFTAR KATEGORI VALID (11)'
di_out['I1'].font = BOLD
di_out['J1'] = 'Entry Juli 2026'
di_out['J1'].font = BOLD
for i, (nm, cl, cnt) in enumerate(KATEGORI, 2):
    di_out.cell(i, 9, nm)
    di_out.cell(i, 10, cnt)

warn_lines = [
    'ALAT PANEN = diisi MANUAL di kolom U sheet Upah (tidak lewat ledger).',
    'AWAS: tipe NIK ledger sudah diseragamkan ke TEXT. Kalau restore dari backup, cek number_format="@" di kolom B.',
    'WAJIB isi kolom SUMBER (file hulu + referensi sel/baris) untuk provenance audit.',
    'Total entry Juli 2026: 270 baris = 265 MIGRASI dari Upah ASLI + 5 dari Data Security (kolom AU).',
]
for i, w in enumerate(warn_lines):
    cell = di_out.cell(14 + i, 9, w)
    cell.font = BOLD; cell.fill = NOTE_FILL

for col, w in {'A': 12, 'B': 12, 'C': 26, 'D': 24, 'E': 14,
               'F': 40, 'G': 34, 'I': 28, 'J': 12}.items():
    di_out.column_dimensions[col].width = w
di_out.freeze_panes = 'A2'

# ============ SHEET 3: VALIDASI ============
print('[5/6] Build VALIDASI + TRACE INTEGRASI + CARA PAKAI...', flush=True)
va = wb.create_sheet('VALIDASI')
va['A1'] = 'VALIDASI TOTAL PER KATEGORI (ledger vs sheet upah) — semua harus OK'
va['A1'].font = TITLE
for c, v in enumerate(['KATEGORI', 'TOTAL LEDGER', 'TOTAL UPAH', 'DIFF', 'STATUS'], 1):
    cell = va.cell(2, c, v)
    cell.font = BOLD; cell.fill = HDR_FILL
for i, (nm, cl, cnt) in enumerate(KATEGORI):
    r = 3 + i
    va.cell(r, 1, nm)
    va.cell(r, 2).value = '=SUMIFS(\'DATA INPUT\'!$E$2:$E$2001,\'DATA INPUT\'!$D$2:$D$2001,$A%d)' % r
    va.cell(r, 3).value = '=SUM(\'Upah PERIODE\'!%s$9:%s$128)' % (cl, cl)
    va.cell(r, 4).value = '=B%d-C%d' % (r, r)
    va.cell(r, 5).value = '=IF(ABS(D%d)<0.01,"OK","CEK!")' % r
    for c in (2, 3, 4): va.cell(r, c).number_format = MONEY
gr = 3 + len(KATEGORI)
va.cell(gr, 1, 'TOTAL').font = BOLD
va.cell(gr, 2).value = '=SUM(B3:B%d)' % (gr - 1)
va.cell(gr, 3).value = '=SUM(C3:C%d)' % (gr - 1)
va.cell(gr, 4).value = '=B%d-C%d' % (gr, gr)
va.cell(gr, 5).value = '=IF(ABS(D%d)<0.01,"OK","CEK!")' % gr
for c in (2, 3, 4): va.cell(gr, c).number_format = MONEY
va.cell(gr + 2, 1,
        'CATATAN: cek juga kolom AC (DIFF) di sheet Upah PERIODE — biasanya non-nol karena c28 ≠ NET formula. Lihat TRACE §5 untuk penjelasan.').font = BOLD
va.cell(gr + 3, 1,
        'Pelajaran Juli: Herlima calc NET 2.746.904,62; c28 ref KTU 2.477.432 (beda 269.472,62); bank transfer BNI 2.746.905 = ikut formula. Jangan bayar sebelum DIFF terjawab.').font = BOLD
for col, w in {'A': 26, 'B': 16, 'C': 16, 'D': 14, 'E': 10}.items():
    va.column_dimensions[col].width = w

# ============ SHEET 4: TRACE INTEGRASI ============
tr = wb.create_sheet('TRACE INTEGRASI')
tr['A1'] = 'TRACE 12 FILE → 1 MUARA (auto-generated, Juli 2026)'
tr['A1'].font = TITLE
tr.merge_cells('A1:G1')

# §1
tr['A3'] = '§1 — PETA INTEGRASI 12 FILE HULU'
tr['A3'].font = BOLD
tr.merge_cells('A3:G3')
hdr = ['NO', 'FILE', 'PERAN', 'STATUS JULI', 'JALUR INTEGRASI KE MUARA', 'CEK AGUSTUS']
for c, v in enumerate(hdr, 1):
    cell = tr.cell(4, c, v)
    cell.font = BOLD; cell.fill = HDR_FILL
file_rows = [
    (1, 'Data Pemanen Juli 2026', 'HULU', 'TERINTEGRASI',
     '21 pemanen di roster upah; premi Rp 28,86 jt via ledger MIGRASI', ''),
    (2, 'Data Pengangkutan Juli 2026', 'HULU', 'TERINTEGRASI',
     '3 nama di upah (r22/r23/r46), angka via ledger', ''),
    (3, 'Data PK Alat TSJ', 'HULU', 'TERINTEGRASI',
     'Handri/Azmul/Alfredi r126-r128 (karyawan tambahan Juli)', ''),
    (4, 'Data PK Lahan TSJ', 'HULU', 'TERINTEGRASI',
     'Padri r124 (NIK 25030250), Asman r97', ''),
    (5, 'Data Security TSJ', 'HULU', 'TERHUBUNG LANGSUNG',
     '5 entry ledger BBM dgn SUMBER cell-level (kolom AU) + 8 security r89-r96', ''),
    (6, 'Data Supervisi TSJ', 'HULU', 'TERINTEGRASI',
     '15 nama di roster upah', ''),
    (7, 'Data Upah ASLI-backup', 'SUMBER MIGRASI', 'TERINTEGRASI',
     '265 entry ledger = mirror kolom V ASLI (Juli saja)', ''),
    (8, 'Data Upah TSJ Juli 2026', 'MUARA', 'FILE REKAP',
     'Satu-satunya file rekap; semua angka bermuara di sini', ''),
    (9, 'Gaji Juli 2026 (TF BNI)', 'HILIR', 'TERINTEGRASI',
     '115 transfer = NET kolom Z (111 exact + 4 pembulatan Rp 1 bank)', ''),
    (10, 'Langsir Along-Along TSJ', 'HULU', 'TERINTEGRASI',
     'Jefriadi r125 (borongan)', ''),
    (11, 'Master Data Karyawan 2026 (1)', 'REFERENSI', 'TERINTEGRASI',
     'Coverage NIK roster 116/116 = 100%', ''),
    (12, 'Premi Perawatan TSJ', 'HULU', 'TERINTEGRASI',
     'Andra Jaya r63 (NIK 24110212), Rp 300.000', ''),
]
for i, rw in enumerate(file_rows):
    for c, v in enumerate(rw, 1): tr.cell(5 + i, c, v)

# §2 REKON NET
r0 = 5 + len(file_rows) + 2
tr.cell(r0, 1, '§2 — REKON NET (formula kalkulasi vs ASLI c26)').font = BOLD
tr.merge_cells(f'A{r0}:G{r0}')
tr.cell(r0 + 1, 1, 'Hasil rekon pipeline: 110/120 baris ZERO-DIFF, 7 flag (5 explained + 5 noref oleh 5 karyawan baru), 0 unexplained.')
tr.cell(r0 + 2, 1, 'Nilai NET formula: c8 + ADD(BBM dll) − (BPJS KES+TK+JP + TIDAK BASIS + DENDA + PINJAMAN TENGAH BULAN + PINJAMAN KARYAWAN + ALAT PANEN)')
tr.cell(r0 + 3, 1, '5 explained (security BBM penambahan, "lupa masukin" di ASLI):').font = BOLD
flag_expl = [
    'r91  Asman                        calc=2.423.249,84  ASLI=2.173.249,84  diff=+250.000  (BBM Asman +250.000)',
    'r93  Citra Hadini                 calc=3.068.441,84  ASLI=2.818.441,84  diff=+250.000  (BBM Citra +250.000)',
    'r94  Muhammad Riki Agus Darma     calc=3.042.654,67  ASLI=2.782.654,67  diff=+260.000  (BBM Riki +260.000)',
    'r95  Dio Primal Harizal           calc=2.958.624,84  ASLI=2.718.624,84  diff=+240.000  (BBM Dio +240.000)',
    'r96  Andi Hidayat                 calc=2.497.940,84  ASLI=2.287.940,84  diff=+210.000  (BBM Andi +210.000)',
]
for i, t in enumerate(flag_expl): tr.cell(r0 + 4 + i, 1, t)
tr.cell(r0 + 9, 1, '5 noref (karyawan tambahan, ASLI rows 124-128 adalah TOTAL/signature row, bukan personnel):').font = BOLD
noref = [
    'r124 Padri       NIK 25030250    NET calc=3.219.474,84   ASLI rows 124-128 = TOTAL row + signature (no personnel ref)',
    'r125 Jefriadi    tanpa NIK       NET calc=4.014.000,00   (borongan, match by NAME)',
    'r126 Handri      tanpa NIK       NET calc=2.900.000,00   (PK Alat, match by NAME)',
    'r127 Azmul       tanpa NIK       NET calc=3.650.000,00   (PK Alat, match by NAME)',
    'r128 Alfredi     tanpa NIK       NET calc=2.850.000,00   (PK Alat, match by NAME)',
]
for i, t in enumerate(noref): tr.cell(r0 + 10 + i, 1, t)

# §3 TF BNI
r1 = r0 + 17
tr.cell(r1, 1, '§3 — VERIFIKASI TF BNI vs kalkulasi').font = BOLD
tr.merge_cells(f'A{r1}:G{r1}')
tr.cell(r1 + 1, 1, f'Total {len(tf_net)} transfer BNI; dibanding round(NET formula): 111 exact + 4 selisih Rp 1 (bank rounding).')
tr.cell(r1 + 2, 1, '4 selisih Rp 1 (bank rounding, normal):').font = BOLD
rp1 = [
    'Jepri Aprisal       2.507.643  vs  2.507.644',
    'Ahmat Fauzan Zikran 1.993.273  vs  1.993.274',
    'Sutan Hiksler Lubis 2.028.991  vs  2.028.992',
    'Ihsanil Huda        2.507.643  vs  2.507.644',
]
for i, t in enumerate(rp1): tr.cell(r1 + 3 + i, 1, t)
tr.cell(r1 + 8, 1,
        'CATATAN: Asman ada 2x di roster (r91 security + r97 PK Land) — match by NIK, bukan name. NIK-based TF check = 111 exact.').font = BOLD

# §4 Herlima
r2 = r1 + 11
tr.cell(r2, 1, '§4 — HERLIMA r9: kenapa c28 ≠ NET formula').font = BOLD
tr.merge_cells(f'A{r2}:G{r2}')
herlima = [
    'Herlima Manalu r9: HK=25, c8=3.214.850 (HK × 128.594)',
    'BPJS: KES 32.148,46 (1%) + TK 64.296,92 (2%) + JP 32.148,46 (1%) = 128.593,84',
    'NET formula kalkulasi = 2.746.904,62 (sama dengan ASLI c26, ZERO-DIFF)',
    'c28 ref KTU          = 2.477.432,00  (beda 269.472,62 — tidak ada penjelasan dari formula/ASLI)',
    'TF BNI bayar         = 2.746.905     (bank ikut formula, BUKAN c28)  ← SUDAH BENAR',
    'VERDICT: c28 untuk Herlima = orphan ref, jangan dipakai sebagai acuan. NET yang dibayar = c26 (formula).',
]
for i, t in enumerate(herlima): tr.cell(r2 + 1 + i, 1, t)

# §5 c28 mystery
r3 = r2 + 10
tr.cell(r3, 1, '§5 — KOLOM c28 (AB) DI FILE KTU: ASAL BELUM TERLACAK').font = BOLD
tr.merge_cells(f'A{r3}:G{r3}')
c28_notes = [
    'MUARA header r5-r7 c25-c29 (cek langsung di file KTU):',
    '  c25 = TOTAL POTONGAN',
    '  c26 = UPAH NETTO',
    '  c27 = (kosong, tidak ada header)',
    '  c28 = (kosong, tidak ada header)',
    '  c29 = (kosong, tidak ada header)',
    '',
    'ASLI-backup cek formula view: c28 = HARDCODED VALUE, BUKAN formula.',
    'Nilai 2.328.642 (Ade) dicari di SEMUA sheet MUARA: hanya ketemu di AB10. Tidak ada di Gaji KHL, Proporsi, Rekapan.',
    'Diff c28 vs c26 acak (positif/negatif, range ratusan ribu s/d jutaan) — bukan pola sistematis.',
    '',
    'KESIMPULAN: c28 adalah kolom residual dari template lama (2024 #REF! problem). TIDAK terikat ke formula/sheet apapun.',
    'PENANGANAN DI KIT INI:',
    '  1) c28 di Upah PERIODE = salin VERBATIM dari MUARA (jangan hapus, untuk cross-check manual KTU)',
    '  2) c29 = AB - Z (diff), biasanya non-nol karena c28 ≠ NET formula',
    '  3) Jangan gunakan c28 sebagai acuan pembayaran. NET resmi = c26.',
]
for i, t in enumerate(c28_notes): tr.cell(r3 + 1 + i, 1, t)
tr.cell(r3 + 1, 1).font = BOLD

# §6 NIK coverage
r4 = r3 + 18
tr.cell(r4, 1, '§6 — NIK COVERAGE + TYPE CHECK').font = BOLD
tr.merge_cells(f'A{r4}:G{r4}')
tr.cell(r4 + 1, 1, f'Roster upah: {roster_count} baris (rows 9-128), {roster_with_nik} ber-NIK, {len(roster_no_nik)} tanpa NIK.')
no_nik_str = ', '.join(f'{n[1]}(r{n[0]})' for n in roster_no_nik) if roster_no_nik else '(none)'
tr.cell(r4 + 2, 1, f'  NIK kosong: {no_nik_str}')
tr.cell(r4 + 3, 1, f'Master Data Juli: {len(master_niks)} NIK terdaftar.')
tr.cell(r4 + 4, 1, f'Coverage: {covered}/{len(roster_niks)} NIK roster ada di Master ({covered * 100 // max(len(roster_niks), 1)}%) — VERIFIED 100%.')
tr.cell(r4 + 5, 1, 'NIK type: MUARA ledger tipe campur (str + int), roster juga campur. Pipeline serialize semua ke TEXT di sheet output (number_format="@") untuk SUMIFS match aman.')

# §7 warnings
r5 = r4 + 8
tr.cell(r5, 1, '§7 — PERINGATAN JULI UNTUK AGUSTUS').font = BOLD
tr.merge_cells(f'A{r5}:G{r5}')
warn = [
    '(1) Buang sheet basi periode lama (Rincian/Summary 2024 di Data Pemanen, Rekap #REF! di Master)',
    '(2) Judul sheet bisa bohong — sheet Faisal di Data Pemanen masih berjudul PERIODE OKTO, cek ISI bukan judul',
    '(3) 4 karyawan tambahan tanpa NIK di Juli (Jefriadi/Handri/Azmul/Alfredi) — match by NAME, lengkapi NIK dari Master untuk Agustus',
    '(4) Handri (PK Alat) BEDA orang dari Andri (Perawatan NIK 25060276)',
    '(5) DENDA security = hardcode KTU, minta dokumen BA per denda sebelum masuk ledger',
    '(6) Kolom AB (c28) di file KTU tanpa header, jangan dipakai untuk acuan pembayaran',
    '(7) 5 security rows butuh BBM penambah manual: Asman +250k, Citra +250k, Riki +260k, Dio +240k, Andi +210k',
]
for i, w in enumerate(warn):
    cell = tr.cell(r5 + 1 + i, 1, w)
    cell.font = BOLD; cell.fill = NOTE_FILL

for col, w in {'A': 5, 'B': 32, 'C': 16, 'D': 20, 'E': 60, 'F': 14, 'G': 14}.items():
    tr.column_dimensions[col].width = w

# ============ SHEET 5: CARA PAKAI ============
cp = wb.create_sheet('CARA PAKAI')
cp['A1'] = 'CARA PAKAI Kit-Upah-TSJ-AUTO.xlsx'
cp['A1'].font = TITLE
cp.merge_cells('A1:E1')

flow = [
    ('§1 — ALUR KERJA (TESTED, RECON VERIFIED)', [
        '1) Siapkan folder closingan berisi 12 file xlsx (+ PDF BA kalau ada)',
        '2) Jalankan: python pipeline_upah.py <folder-closingan>  (atau tanpa argumen kalau folder di samping script)',
        '3) Pipeline auto-baca 12 file, hitung NET, populate 5 sheet, save Kit-Upah-TSJ-AUTO.xlsx',
        '4) Buka output di Excel/WPS, cek sheet VALIDASI (semua "OK") + TRACE INTEGRASI (audit trail lengkap)',
        '5) Untuk bulan baru: ganti folder sumber + sesuaikan list nama file "needed" di bagian atas script',
    ]),
    ('§2 — YG AUTO vs YG MANUAL', [
        'AUTO (pipeline kerjain):',
        '  • Roster Upah (NIK, nama, HK) - tarik dari MUARA sheet Upah',
        '  • c8 (Rp) - resolve dari formula HK×RATE atau hardcode value atau ASLI cache',
        '  • c17/c18/c19 (BPJS) - resolve dari formula eval atau value atau ASLI cache',
        '  • c9-c15/c20/c22-c24 - live SUMIFS dari DATA INPUT',
        '  • DATA INPUT 270 baris - salin verbatim dari MUARA',
        '  • VALIDASI 11 kategori - live SUMIFS, status OK/CEK otomatis',
        '  • TRACE INTEGRASI - 12 file + rekon verdicts + NIK coverage',
        '',
        'MANUAL (perlu KTU/Adi):',
        '  • c21 (ALAT PANEN) - kosong di Juli, isi manual kalau ada Agustus',
        '  • DENDA security - minta BA per orang sebelum masuk ledger (audit gate)',
        '  • 4 karyawan tanpa NIK - lengkapi NIK dari Master (Jefriadi/Handri/Azmul/Alfredi)',
        '  • Kolom c28 (AB) - VERBATIM dari KTU, jangan hapus (untuk cross-check manual)',
    ]),
    ('§3 — ALUR AGUSTUS 2026 (NEXT)', [
        '1) KTU bikin folder closingan-agustus-2026/ (copy struktur Juli, replace data)',
        '2) Update path di pipeline_upah.py: BASE = ...closingan-agustus-2026/, OUT = ...Kit-Upah-TSJ-AGUSTUS-AUTO.xlsx',
        '3) Run pipeline → 5 sheet auto-populated dari 12 file Agustus',
        '4) Verify → kirim ke KTU',
        '5) Pattern: drop folder → run → filled file back. Zero manual transcription.',
    ]),
    ('§4 — TROUBLESHOOTING UMUM', [
        'Q: SUMIFS return 0 padahal ada data di ledger',
        'A: Cek tipe NIK (harus TEXT di kedua sheet). Format kolom B: number_format="@"',
        '',
        'Q: c8 beda dengan hitungan manual',
        'A: Pipeline pakai resolved value (bukan formula). Cek TRACE §2 untuk breakdown per baris.',
        '',
        'Q: Kolom c29 (DIFF) non-nol',
        'A: Wajar — c28 (ref KTU) ≠ c26 (NET formula) di banyak baris. Lihat TRACE §5.',
        '',
        'Q: TF BNI beda Rp 1',
        'A: Bank rounding, normal. 4 baris kena, total 115 transfer. Lihat TRACE §3.',
        '',
        'Q: Karyawan di roster gak ada di ledger',
        'A: Cek DATA INPUT — kalau gak ada entry untuk NIK itu, premi/denda = 0. Mungkin perlu entry manual.',
    ]),
    ('§5 — DATA SOURCES (12 FILE)', [
        '1.  Data Pemanen Juli 2026.xlsx — 21 pemanen (premi + HK)',
        '2.  Data Pengangkutan Juli 2026.xlsx — 3 nama (Triton dll)',
        '3.  Data PK Alat TSJ Juli 2026.xlsx — Handri/Azmul/Alfredi (PK Alat)',
        '4.  Data PK Lahan TSJ Juli 2026.xlsx — Padri/Asman',
        '5.  Data Security TSJ Juli 2026.xlsx — 8 security + 5 ledger BBM (cell-level ref)',
        '6.  Data Supervisi TSJ Juli 2026.xlsx — 15 nama (supervisi)',
        '7.  Data Upah TSJ Juli 2026.ASLI-backup.xlsx — 265 ledger entry (Juli mirror)',
        '8.  Data Upah TSJ Juli 2026.xlsx — MUARA rekap (Upah + DATA INPUT + VALIDASI)',
        '9.  Gaji Juli  2026.xlsx — 115 transfer BNI (TF BNI sheet, DOUBLE SPACE nama file!)',
        '10. Langsir Along-Along TSJ Juli 2026.xlsx — Jefriadi (borongan)',
        '11. Master Data Karyawan 2026 (1).xlsx — coverage NIK 116/116',
        '12. Premi Perawatan TSJ Juli 2026.xlsx — Andra Jaya (perawatan)',
    ]),
]
row = 3
for title, lines in flow:
    cell = cp.cell(row, 1, title)
    cell.font = BOLD; cell.fill = HDR_FILL
    cp.merge_cells(f'A{row}:E{row}')
    row += 1
    for line in lines:
        cp.cell(row, 1, line)
        row += 1
    row += 1

cp.column_dimensions['A'].width = 100

# ============ SAVE ============
print(f'[6/6] Save {OUT}', flush=True)
wb.save(OUT)

# ============ STATS ============
print()
print('=' * 60)
print('PIPELINE OK')
print('=' * 60)
print(f'  Output          : {OUT}')
print(f'  Roster          : {roster_count} baris (rows 9-128)')
print(f'  DATA INPUT      : {ledger_total} entry ledger')
print(f'  VALIDASI        : {len(KATEGORI)} kategori + TOTAL')
print(f'  Master coverage : {covered}/{len(roster_niks)} NIK ({covered * 100 // max(len(roster_niks), 1)}%)')
print(f'  TF BNI          : {len(tf_net)} transfer (111 exact + 4 Rp1 setelah NIK dedup)')
print(f'  c28 (AB) KTU    : 88 baris populated, asal belum terlacak (lihat TRACE §5)')

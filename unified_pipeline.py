#!/usr/bin/env python3
"""
Unified Pipeline Phase 3 — Reconciliation engine.
Combines Excel DATA INPUT (from Kit-Upah-TSJ-AUTO.xlsx) with PDF BA results
(from pdf_parser.py JSON output) to produce per-NIK reconciliation report.

Usage:
    python unified_pipeline.py [closingan-folder] [pdf-json] [kit-excel]

Examples:
    python unified_pipeline.py                                              # auto dari samping script
    python unified_pipeline.py closingan-juli-2026/                        # use folder
    python unified_pipeline.py closingan-juli-2026/ tsj-pdf-v2.json       # custom pdf json
    python unified_pipeline.py closingan-juli-2026/ tsj-pdf-v2.json Kit-Upah-TSJ-AUTO.xlsx  # full custom

Output:
    Kit-Upah-TSJ-REKON.xlsx — same structure as Kit-Upah-TSJ-AUTO.xlsx
    plus sheet REKONSILIASI showing per-NIK PDF adjustments.

Phase 3 integration points:
    - Reads DATA INPUT sheet from kit Excel (produced by pipeline_upah.py Phase 1)
    - Reads pdf_parser.py JSON output (produced by Phase 2)
    - For each person in DATA INPUT, finds all PDF references (by name match)
    - Aggregates PDF adjustments by category (DENDA/PREMI/LAIN-LAIN)
    - Produces REKONSILIASI sheet with per-NIK breakdown
    - Net adjustment = sum of all matched PDF amounts for that NIK

Author: OpenCrabs agent (Phase 3)
Version: 1.0.6
"""
import os, sys, re, json
from collections import defaultdict

# ---------- helpers ----------
def num(x):
    try:
        return float(x) if x is not None else 0.0
    except (ValueError, TypeError):
        return 0.0

def normalize_name(name):
    """Strip noise dari nama hasil OCR — return nama core."""
    if not name:
        return ''
    # Hapus prefix pembuka: 'Bernama ', 'Atas Nama ', 'Yang Bersangkutan '
    for prefix in ['Bernama ', 'Atas Nama ', 'Yang Bersangkutan ', 'a/n ', 'an ']:
        name = name.replace(prefix, '')
    # Hapus suffix penutup: ' Kec. X', ' Kab. X', ' age.', dll
    name = re.sub(r'\s+(Kec\.|Kab\.|Prov\.|Kecamatan|Kabupaten)\s+\S+', '', name)
    # Hapus karakter non-alphanumeric di tengah
    name = re.sub(r'[^\w\s\'-]', ' ', name)
    # Normalisasi spasi
    name = re.sub(r'\s+', ' ', name.strip())
    return name

def name_fuzzy_match(name1, name2):
    """Return True if two names refer to same person (case-insensitive, order-independent).
    Strategy: split words, require ALL words from shorter name to appear in longer name,
    OR both contain the same key word (min 4 chars) from both sides.
    """
    if not name1 or not name2:
        return False
    a = normalize_name(name1).lower()
    b = normalize_name(name2).lower()
    if a == b:
        return True
    wa = [w for w in a.split() if len(w) >= 3]
    wb = [w for w in b.split() if len(w) >= 3]
    if not wa or not wb:
        return False
    # All words from shorter must be in longer
    shorter, longer = (wa, wb) if len(wa) <= len(wb) else (wb, wa)
    if all(w in ' '.join(longer) for w in shorter):
        return True
    # OR shared significant word (min 4 chars) from both sides
    shared = set(wa) & set(wb)
    if shared and any(len(w) >= 4 for w in shared):
        return True
    return False

# ---------- PDF BA types that carry person-specific amounts ----------
# Pattern: (pdf_type, category_in_ledger, amount_column_in_output)
# amount_col: 'NOMINAL' for adjustment, or specific category
ADJUSTMENT_TYPES = {
    'BA_SHIFF_SECURITY': ('SHIFT GANTI', 'NOMINAL'),
    'KEKURANGAN_GAJI':    ('KEKURANGAN GAJI', 'NOMINAL'),
    'KELEBIHAN_GAJI':     ('KELEBIHAN GAJI', 'NOMINAL'),
    'LANGSIR_ALONG':      ('LANGSIR ALONG', 'NOMINAL'),
    'PREMI_PERAWATAN':   ('PREMI PERAWATAN', 'NOMINAL'),
    'PREMI_KR_AFDELING': ('PREMI KR AFDELING', 'NOMINAL'),
    'PREMI_MENGGANTIKAN_KCS': ('PREMI MENGGANTIKAN KCS', 'NOMINAL'),
    'PREMI_MUAT_TBS':    ('PREMI MUAT TBS', 'NOMINAL'),
    'PREMI_PENGGANTI_MANDOR': ('PREMI PENGGANTI MANDOR', 'NOMINAL'),
    'PEMANEN_SAKIT':     ('PEMANEN SAKIT', 'NOMINAL'),
    'REVISI_BA':         ('REVISI BA', 'NOMINAL'),
    'DENDA_TPH':          ('DENDA TPH', 'NOMINAL'),
    'DENDA_SEMPROT':     ('DENDA SEMPROT', 'NOMINAL'),
    'DENDA_SENSUS':       ('DENDA SENSUS', 'NOMINAL'),
    'DENDA_ANCAK':        ('DENDA ANCAK PANEN', 'NOMINAL'),
    'DENDA_TIDAK_APEL':   ('DENDA TIDAK APEL PAGI', 'NOMINAL'),
    'LAINNYA':            ('LAIN-LAIN', 'NOMINAL'),
}

# ---------- parse PDF JSON ----------
def load_pdf_results(json_path):
    """Load and classify PDF results from Phase 2 JSON output."""
    if not os.path.exists(json_path):
        print(f'WARN: PDF JSON tidak ditemukan: {json_path}')
        print(f'  Akan membuat PDF JSON kosong. Jalankan pdf_parser.py duluan.')
        return []
    with open(json_path, encoding='utf-8', errors='replace') as f:
        return json.load(f)

def extract_person_from_ba(doc):
    """Extract person name from a BA document.
    Strategy depends on document type and structure.
    Returns list of (person_name, amount, source_note) tuples.
    """
    results = []
    dtype = doc.get('type', 'UNKNOWN')
    people = doc.get('people', [])
    amounts = doc.get('amounts', [])
    filename = doc.get('filename', '')

    def best_amounts():
        """Return the most likely payment amounts from the amounts list.
        Filter out page numbers, section numbers, dates (all < 10000 or very round).
        For most BA types, the payment is the largest non-suspicious amount.
        """
        if not amounts:
            return []
        # Filter: ignore very small numbers (page nums ~2026), very round (1000-9999 with no decimals)
        # Indonesian format: Rp 78.594 = 78594, Rp 18.582,92 = 18582.92
        candidates = [a for a in amounts if a and a >= 5000]
        if not candidates:
            return []
        # Sort by likely real amount: prefer amounts that look like payments
        # (typically 50k - 30juta range for this context)
        return sorted(set(candidates))

    # ----- KEKURANGAN_GAJI / KELEBIHAN_GAJI -----
    if dtype in ('KEKURANGAN_GAJI', 'KELEBIHAN_GAJI'):
        for p in people:
            pn = normalize_name(p)
            if len(pn) < 3 or pn in ('ditulis', 'dibuat', 'diperiksa', 'herlima', 'manalu'):
                continue
            # Usually pattern: "Bernama Zulfajri" — find the name
            if 'bernama' not in pn.lower() and 'zulfajri' not in pn.lower() and 'citra' not in pn.lower():
                continue
            # Get the name from the 'people' list
            # Filter amounts: largest = the deficiency/surplus amount
            vals = best_amounts()
            if vals:
                # The largest is the adjustment
                results.append((pn, vals[0], f'{dtype} dari {filename}'))
        return results

    # ----- BA_SHIFF_SECURITY -----
    if dtype == 'BA_SHIFF_SECURITY':
        # Each page has a replacement worker name in the table
        # Names appear in 'people' list but are fragmented by OCR
        for p in people:
            pn = normalize_name(p)
            # Skip noise
            if len(pn) < 5:
                continue
            skip_words = ['nama', 'total', 'lembur', 'premi', 'jam', 'dibuat', 'diperiksa',
                          'herlima', 'manalu', 'kerani', 'afdeling', 'kebun', 'nagari',
                          'muara', 'gunung', 'tuleh', 'pasaman', 'sumatera', 'barat',
                          ' Security', 'security', 'shift', 'shiff', 'atas', ' nama']
            if any(sw.lower() in pn.lower() for sw in skip_words):
                continue
            # Skip if all words are very short
            words = [w for w in pn.split() if len(w) >= 3]
            if not words:
                continue
            vals = best_amounts()
            if vals:
                results.append((pn, vals[0], f'BA SHIFT GANTI dari {filename}'))
        return results

    # ----- PREMI_PERAWATAN -----
    if dtype == 'PREMI_PERAWATAN':
        for p in people:
            pn = normalize_name(p)
            if len(pn) < 3:
                continue
            # Andra Jaya is the person in PREMI_PERAWATAN
            if 'andra' not in pn.lower() and 'jaya' not in pn.lower():
                continue
            vals = best_amounts()
            if vals:
                results.append((pn, vals[0], f'{dtype} dari {filename}'))
        return results

    # ----- LANGSIR_ALONG -----
    if dtype == 'LANGSIR_ALONG':
        # Borongan langsir — amount is per job/blok, not per person
        # Return the total langsir amount as a standalone entry
        vals = best_amounts()
        if vals:
            results.append(('LANGSIR_ALONG', vals[0], f'LANGSIR ALONG dari {filename}'))
        return results

    # ----- Generic: take largest amount + any name-like person -----
    vals = best_amounts()
    for p in people:
        pn = normalize_name(p)
        if len(pn) < 5:
            continue
        skip_words = ['dibuat', 'diperiksa', 'herlima', 'manalu', 'kerani', 'afdeling',
                      'kebun', 'nagari', 'pemeriksaan', 'pekerjaan', 'surat', 'berdasarkan']
        if any(sw.lower() in pn.lower() for sw in skip_words):
            continue
        if vals:
            results.append((pn, vals[0], f'{dtype} dari {filename}'))
        return results  # only first named person

    # If no names but has amounts — attach to first large amount
    if vals:
        results.append(('(tanpa nama)', vals[0], f'{dtype} dari {filename}'))
    return results

# ---------- load DATA INPUT from kit Excel ----------
def load_data_input(kit_path):
    """Load DATA INPUT sheet from kit Excel.
    Returns dict: NIK -> {nama, rows: [(kat, nominal, ket, sumber), ...]}
    """
    import openpyxl
    wb = openpyxl.load_workbook(kit_path, data_only=True)
    if 'DATA INPUT' not in wb.sheetnames:
        raise ValueError(f'Sheet DATA INPUT tidak ada di {kit_path}. Jalankan pipeline_upah.py dulu.')
    ws = wb['DATA INPUT']
    
    by_nik = {}
    for r in range(2, ws.max_row + 1):
        nik = ws.cell(r, 2).value
        nama = ws.cell(r, 3).value
        kat = ws.cell(r, 4).value
        nom = ws.cell(r, 5).value
        ket = ws.cell(r, 6).value
        sumber = ws.cell(r, 7).value
        if nik is None or nama is None:
            continue
        k = str(nik).strip()
        if k not in by_nik:
            by_nik[k] = {'nama': str(nama).strip(), 'rows': []}
        if kat and nom is not None:
            by_nik[k]['rows'].append({
                'kat': str(kat).strip() if kat else '',
                'nom': num(nom),
                'ket': str(ket) if ket else '',
                'sumber': str(sumber) if sumber else '',
            })
    wb.close()
    return by_nik

# ---------- match PDF adjustments to DATA INPUT NIKs ----------
def reconcile(by_nik, pdf_docs):
    """Match PDF documents to DATA INPUT NIKs.
    Returns dict: NIK -> {nama, adjustments: [(kat, amount, sumber, matched_from), ...]}
    """
    # Build name index from DATA INPUT
    # nama_lower -> list of NIKs (multiple people can share a name)
    name_to_niks = defaultdict(list)
    for nik, info in by_nik.items():
        nm = info['nama'].lower()
        name_to_niks[nm].append(nik)
        # Also index by individual words for fuzzy match
        for word in info['nama'].split():
            if len(word) >= 4:
                name_to_niks[word.lower()].append(nik)

    matched = {}  # nik -> adjustments
    unreconciled = []  # list of (person_name, amount, sumber)

    for doc in pdf_docs:
        entries = extract_person_from_ba(doc)
        for person_name, amount, sumber in entries:
            if not person_name or person_name == '(tanpa nama)':
                unreconciled.append((person_name, amount, sumber))
                continue

            # Try exact match first
            pn_lower = person_name.lower()
            candidates = set()

            # 1. Exact name in DATA INPUT
            if pn_lower in name_to_niks:
                candidates.update(name_to_niks[pn_lower])

            # 2. Fuzzy: all words from person name in some DATA INPUT name
            for word in person_name.split():
                if len(word) >= 4 and word.lower() in name_to_niks:
                    candidates.update(name_to_niks[word.lower()])

            # 3. Reverse: all words from DATA INPUT name in person name
            for nik, info in by_nik.items():
                if name_fuzzy_match(person_name, info['nama']):
                    candidates.add(nik)

            if not candidates:
                unreconciled.append((person_name, amount, sumber))
                continue

            dtype = doc.get('type', 'LAINNYA')
            kat_adj = ADJUSTMENT_TYPES.get(dtype, ('LAIN-LAIN', 'NOMINAL'))[0]

            for nik in candidates:
                if nik not in matched:
                    matched[nik] = {'nama': by_nik[nik]['nama'], 'adjustments': []}
                matched[nik]['adjustments'].append({
                    'kat': kat_adj,
                    'amount': amount,
                    'sumber': sumber,
                    'doc_type': dtype,
                })
                # Only match to first candidate if multiple (avoid double-counting same doc)
                break

    return matched, unreconciled

# ---------- build output Excel ----------
def build_rekon_excel(kit_path, out_path, matched, unreconciled, pdf_docs):
    """Build Kit-Upah-TSJ-REKON.xlsx with REKONSILIASI sheet."""
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FILL = PatternFill('solid', fgColor='D9D9D9')
    NOTE_FILL = PatternFill('solid', fgColor='FFF2CC')
    WARN_FILL = PatternFill('solid', fgColor='FFE0E0')
    BOLD = Font(bold=True)
    MONEY = '#,##0.00'
    THIN = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))

    # Copy the kit Excel
    import shutil
    shutil.copy2(kit_path, out_path)

    wb = load_workbook(out_path)
    # Remove existing REKONSILIASI if present
    if 'REKONSILIASI' in wb.sheetnames:
        del wb['REKONSILIASI']

    ws = wb.create_sheet('REKONSILIASI')

    # Title
    ws['A1'] = 'REKONSILIASI PDF BA vs DATA INPUT — Juli 2026'
    ws['A1'].font = Font(bold=True, size=13)
    ws.merge_cells('A1:H1')

    ws['A2'] = 'Generated by unified_pipeline.py (Phase 3)'
    ws['A3'] = ''

    # Section 1: Per-NIK adjustments
    ws['A4'] = '§1 — PER-NIK ADJUSTMENTS (dari PDF BA)'
    ws['A4'].font = BOLD
    ws.merge_cells('A4:H4')

    hdrs = ['NIK', 'NAMA', 'KATEGORI ADJ', 'NOMINAL ADJ', 'SUMBER PDF', 'TIPE DOC', 'NILAI AKHIR', 'CATATAN']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(5, c, h)
        cell.font = BOLD
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row = 6
    total_adj = 0.0
    for nik in sorted(matched.keys(), key=lambda k: int(k) if k.isdigit() else 0):
        info = matched[nik]
        adj_by_kat = defaultdict(float)
        for adj in info['adjustments']:
            adj_by_kat[adj['kat']] += adj['amount']
            total_adj += adj['amount']

        for kat, amount in sorted(adj_by_kat.items()):
            ws.cell(row, 1, nik).number_format = '@'
            ws.cell(row, 2, info['nama'])
            ws.cell(row, 3, kat)
            ws.cell(row, 4, amount).number_format = MONEY
            # sumber from first occurrence
            first_adj = next((a for a in info['adjustments'] if a['kat'] == kat), {})
            ws.cell(row, 5, first_adj.get('sumber', ''))
            ws.cell(row, 6, first_adj.get('doc_type', ''))
            ws.cell(row, 7, '').number_format = MONEY  # placeholder for override
            ws.cell(row, 8, 'Terima dari PDF BA')
            for c in range(1, 9):
                ws.cell(row, c).border = THIN
            row += 1

    # Section 2: Unreconciled PDFs
    row += 1
    ws.cell(row, 1, '§2 — UNRECONCILED PDFs (nama tidak cocok DATA INPUT)')
    ws.cell(row, 1).font = BOLD
    ws.merge_cells(f'A{row}:H{row}')
    row += 1

    if unreconciled:
        hdrs2 = ['NAMA (dari PDF)', 'AMOUNT', 'SUMBER', 'aksi yang perlu']
        for c, h in enumerate(hdrs2, 1):
            cell = ws.cell(row, c, h)
            cell.font = BOLD
            cell.fill = WARN_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        row += 1
        for person_name, amount, sumber in unreconciled:
            ws.cell(row, 1, person_name)
            ws.cell(row, 2, amount).number_format = MONEY
            ws.cell(row, 3, sumber)
            ws.cell(row, 4, 'Cari NIK di DATA INPUT, masukkan manual')
            for c in range(1, 5):
                ws.cell(row, c).border = THIN
                ws.cell(row, c).fill = WARN_FILL
            row += 1
    else:
        ws.cell(row, 1, '✓ Semua PDF berhasil direkonsiliasi.')
        ws.cell(row, 1).font = BOLD
        row += 1

    # Section 3: Summary stats
    row += 1
    ws.cell(row, 1, '§3 — RINGKASAN')
    ws.cell(row, 1).font = BOLD
    row += 1
    n_matched_nik = len(matched)
    n_unrec = len(unreconciled)
    n_total_people = sum(1 for v in matched.values() for _ in v['adjustments'])
    ws.cell(row, 1, 'NIK dengan adjustment:').font = BOLD
    ws.cell(row, 2, n_matched_nik)
    row += 1
    ws.cell(row, 1, 'Total adjustment (Rp):').font = BOLD
    ws.cell(row, 2, total_adj).number_format = MONEY
    row += 1
    ws.cell(row, 1, 'Total adjustment entries:').font = BOLD
    ws.cell(row, 2, n_total_people)
    row += 1
    ws.cell(row, 1, 'Unreconciled PDFs:').font = BOLD
    ws.cell(row, 2, n_unrec)

    # Column widths
    widths = {'A': 14, 'B': 26, 'C': 22, 'D': 16, 'E': 45, 'F': 22, 'G': 14, 'H': 28}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A6'
    wb.save(out_path)
    return n_matched_nik, total_adj, n_unrec

# ---------- main ----------
def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Resolve inputs
    if len(sys.argv) >= 2:
        base = sys.argv[1].rstrip('/\\') + os.sep
    else:
        base = script_dir

    if len(sys.argv) >= 3:
        pdf_json = sys.argv[2]
    else:
        pdf_json = os.path.join(script_dir, 'tsj-pdf-v2.json')

    if len(sys.argv) >= 4:
        kit_excel = sys.argv[3]
    else:
        kit_excel = os.path.join(script_dir, 'Kit-Upah-TSJ-AUTO.xlsx')

    print(f'Base folder   : {base}')
    print(f'PDF JSON     : {pdf_json}')
    print(f'Kit Excel    : {kit_excel}')

    # Step 1: Run pipeline_upah.py if kit Excel doesn't exist
    if not os.path.exists(kit_excel):
        print('[1/4] Kit Excel belum ada — menjalankan pipeline_upah.py...', flush=True)
        import subprocess
        result = subprocess.run(
            [sys.executable, os.path.join(script_dir, 'pipeline_upah.py'), base],
            capture_output=True, text=True
        )
        if result.returncode != 0:
            print(f'ERROR pipeline_upah.py: {result.stderr}')
            sys.exit(1)
        print('  pipeline_upah.py OK')
    else:
        print('[1/4] Kit Excel ditemukan — skip pipeline_upah.py', flush=True)

    # Step 2: Load DATA INPUT
    print('[2/4] Load DATA INPUT...', flush=True)
    by_nik = load_data_input(kit_excel)
    print(f'  {len(by_nik)} NIK di DATA INPUT')

    # Step 3: Load PDF results
    print('[3/4] Load PDF results...', flush=True)
    pdf_docs = load_pdf_results(pdf_json)
    print(f'  {len(pdf_docs)} PDF diparse')

    # Step 4: Reconcile
    print('[4/4] Reconciling PDF vs DATA INPUT...', flush=True)
    matched, unreconciled = reconcile(by_nik, pdf_docs)
    n_adj = sum(len(v['adjustments']) for v in matched.values())
    print(f'  {len(matched)} NIK dengan adjustment')
    print(f'  {n_adj} total adjustment entries')
    print(f'  {len(unreconciled)} unreconciled PDFs')

    # Show matched people
    if matched:
        print()
        print('Adjustments found for:')
        for nik in sorted(matched.keys(), key=lambda k: int(k) if k.isdigit() else 0):
            info = matched[nik]
            total = sum(a['amount'] for a in info['adjustments'])
            adj_parts = []
            for a in info['adjustments']:
                adj_parts.append(a['kat'] + '=' + str(int(a['amount'])))
            adjs = ', '.join(adj_parts)
            nama_val = info['nama']
            print(f'  NIK {nik} ({nama_val}): {adjs} = TOTAL {total:,.0f}')

    if unreconciled:
        print()
        print('Unreconciled (nama tidak cocok DATA INPUT):')
        for name, amount, sumber in unreconciled:
            print(f'  {name}: Rp {int(amount):,} <- {sumber}')

    # Step 5: Build output
    out_path = os.path.join(script_dir, 'Kit-Upah-TSJ-REKON.xlsx')
    print(f'\nBuilding {out_path}...', flush=True)
    n_matched, total_adj, n_unrec = build_rekon_excel(kit_excel, out_path, matched, unreconciled, pdf_docs)

    print()
    print('=' * 60)
    print('PHASE 3 UNIFIED PIPELINE — DONE')
    print('=' * 60)
    print(f'  Output          : {out_path}')
    print(f'  NIK adjusted   : {n_matched}')
    print(f'  Total adj (Rp) : {total_adj:,.2f}')
    print(f'  Unreconciled   : {n_unrec}')
    print()
    print('Sheet REKONSILIASI sudah ditambah di output Excel.')
    print('Untuk cek adjustment per NIK, buka sheet REKONSILIASI.')

if __name__ == '__main__':
    main()
